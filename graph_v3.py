import os
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.cm as cm
import numpy as np
from matplotlib.dates import DateFormatter, HourLocator

# 엑셀 파일 로드
wb = load_workbook("/home/ubuntu/watcher/outputs/hw12.xlsx")
output_folder = "./graphs_hw12"

# 그래프 출력 폴더 생성
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 데이터 저장용 딕셔너리 (학생 → 코드 → [(시간, 크기)])
data = {}

# 모든 시트(학생)에서 데이터 읽기
for sheet in wb.sheetnames:
    ws = wb[sheet]  # 현재 시트 (학생)
    student = sheet
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # 헤더 제외
        code_dir, snapshot_file, file_size, file_mtime, timestamp = row
        
        # 데이터 저장 구조: {학생: {코드: [(시간, 크기)]}}
        data.setdefault(student, {}).setdefault(code_dir, []).append((timestamp, file_size))

# 그래프 그리기
for student, code_dict in data.items():
    plt.figure(figsize=(30, 8))  # 가로 크기를 15에서 20으로 증가
    
    # 코드 파일별 색상 지정 (최대 20개 색상)
    cmap = plt.colormaps["tab20"]
    colors = cmap(np.linspace(0, 1, len(code_dict)))
    
    for idx, (code, snapshots) in enumerate(code_dict.items()):
        snapshots.sort()
        times, file_sizes = zip(*snapshots)
        
        # 타임스탬프를 datetime 객체로 변환
        times = [datetime.fromtimestamp(t) for t in times]
        
        plt.plot(times, file_sizes, marker='o', markersize=4, 
                label=os.path.basename(code), color=colors[idx], 
                linewidth=2, linestyle='-')

    # x축 설정 변경
    plt.gca().xaxis.set_major_formatter(DateFormatter('%m/%d\n%H:%M'))  # 월/일\n시간:분 형식
    plt.xticks(times, rotation=45, ha='right')  # 실제 타임스탬프를 눈금으로 사용
    
    plt.xlabel("Time", fontsize=12)
    plt.ylabel("File Size (bytes)", fontsize=12)
    plt.title(f"HW12 - {student}", fontsize=14, pad=20)
    
    # 범례 설정
    plt.legend(title="Code Files", loc='center left', 
              bbox_to_anchor=(1, 0.5),  # 그래프 오른쪽에 범례 위치
              fontsize=10, title_fontsize=12)
    
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.tight_layout()  # 레이아웃 자동 조정

    # 그래프 저장
    output_filename = os.path.join(output_folder, f"{student}_hw12.png")
    plt.savefig(output_filename, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"저장된 그래프: {output_filename}")
