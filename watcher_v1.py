import os
from openpyxl import Workbook
from datetime import datetime

# 최상위 디렉터리 경로 (학번 디렉터리들이 포함된 디렉터리)
root_directory = "/home/ubuntu/Couch"
output_excel = "file_snapshots.xlsx"

# 워크북 생성
wb = Workbook()

# 기본 "Snapshots" 시트 제거 (빈 워크북에서 기본 생성된 시트)
del wb["Sheet"]

# 학번 디렉터리 탐색
for student_dir in os.listdir(root_directory):
    student_path = os.path.join(root_directory, student_dir)
    if not os.path.isdir(student_path):
        continue  # 디렉터리가 아니면 건너뜀

    # 과제 디렉터리 탐색
    for hw_dir in os.listdir(student_path):
        hw_path = os.path.join(student_path, hw_dir)
        if not os.path.isdir(hw_path):
            continue  # 과제 디렉터리가 아니면 건너뜀

        # 각 과제에 대해 시트가 없으면 생성
        if hw_dir not in wb.sheetnames:
            ws = wb.create_sheet(title=hw_dir)  # 과제 이름을 시트 제목으로 사용
            ws.append(["학번", "과제 디렉터리", "과제 코드 디렉터리 (.c)", "스냅샷 파일명", "파일 크기 (bytes)", "마지막 수정 시간"])  # 헤더 추가
        else:
            ws = wb[hw_dir]  # 이미 시트가 존재하면 해당 시트를 사용

        # .c 과제 코드 디렉터리 탐색
        for code_dir in os.listdir(hw_path):
            if not code_dir.endswith(".c"):
                continue  # ".c"로 끝나지 않으면 건너뜀
            code_path = os.path.join(hw_path, code_dir)
            if not os.path.isdir(code_path):
                continue

            # 타임스탬프 스냅샷 파일 탐색
            for snapshot_file in os.listdir(code_path):
                snapshot_path = os.path.join(code_path, snapshot_file)
                if not os.path.isfile(snapshot_path):
                    continue  # 파일이 아니면 건너뜀

                # 파일 정보 수집
                file_size = os.path.getsize(snapshot_path)  # 파일 크기
                file_mtime = datetime.fromtimestamp(os.path.getmtime(snapshot_path)).strftime('%Y-%m-%d %H:%M:%S')  # 마지막 수정 시간

                # 엑셀에 데이터 추가
                ws.append([student_dir, hw_dir, code_dir, snapshot_file, file_size, file_mtime])

# 엑셀 파일 저장
wb.save(output_excel)
print(f"엑셀 파일이 생성되었습니다: {output_excel}")
