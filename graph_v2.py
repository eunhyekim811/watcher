import os
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.cm as cm
import numpy as np

# 엑셀 파일 로드
wb = load_workbook("file_snapshots.xlsx")
output_folder = "./graphs"

# 그래프 출력 폴더 생성
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 데이터 저장용 딕셔너리 (학생 → 과제 → 코드 → [(시간, 크기)])
data = {}

# 모든 시트에서 데이터 읽기
for sheet in wb.sheetnames:
    ws = wb[sheet]  # 현재 시트
    hw_dir = sheet  # 시트명이 과제명
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        student, hw_dir, code_dir, snapshot_file, file_size, file_mtime = row  # 6개의 값 추출


        # 파일명에서 숫자(타임스탬프)만 추출
        match = re.search(r'\d+', snapshot_file)  # 정규표현식으로 숫자 찾기
        if not match:
            continue  # 숫자가 없으면 건너뜀

        try:
            timestamp = datetime.fromtimestamp(int(match.group()))  # UNIX 타임스탬프로 변환
        except ValueError:
            continue  # 변환 실패 시 건너뜀

        # 데이터 저장 구조: {학생: {과제: {코드: [(시간, 크기)]}}}
        data.setdefault(student, {}).setdefault(hw_dir, {}).setdefault(code_dir, []).append((timestamp, file_size))

# 그래프 그리기
for student, hw_dict in data.items():
    for hw, code_dict in hw_dict.items():
        plt.figure(figsize=(50, 30))
        # cmap = plt.cm.get_cmap("tab10", len(code_dict))  # 코드 파일별 색상 지정
        cmap = plt.colormaps["tab20"]
        colors = cmap(np.arange(len(code_dict)))
        
        for idx, (code, snapshots) in enumerate(code_dict.items()):
            # 시간 순 정렬
            snapshots.sort()
            times, file_sizes = zip(*snapshots)

            # 시간 포맷을 연도-월-일-시간-분-초로 변경
            times = [time.strftime("%Y-%m-%d %H:%M:%S") for time in times]
            
            # 그래프 플롯 (작은 마커, 투명도 적용, 점선 스타일)
            plt.plot(times, file_sizes, marker='o', markersize=10, label=code, color=colors[idx])

        # 그래프 설정
        plt.xlabel("Time")
        plt.ylabel("File Size (bytes)")
        plt.title(f"{hw} - {student}")
        plt.xticks(rotation=45)  # x축 레이블 회전
        plt.legend(title="Code Files", loc='upper left', fontsize='large', frameon=True)
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.tight_layout()

        # 그래프 저장
        output_filename = os.path.join(output_folder, f"{student}_{hw}.png")
        plt.savefig(output_filename, dpi=300)
        plt.close()
        print(f"Saved graph: {output_filename}")
