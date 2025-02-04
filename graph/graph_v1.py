import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from datetime import datetime

# 엑셀 파일 로드
wb = load_workbook("file_snapshots.xlsx")
output_folder = "./graphs"

if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 과제별로 그래프를 그리기
for sheet in wb.sheetnames:
    ws = wb[sheet]
    
    # 과제명
    hw_dir = sheet
    
    # 학생별로 그래프 그리기
    for student_dir in set(row[0] for row in ws.iter_rows(min_row=2, values_only=True)):  # 학생별로
        # 시간과 파일 크기 데이터 초기화
        times = []
        file_sizes = []
        
        # 각 행을 읽어서 시간과 파일 크기 추출
        for row in ws.iter_rows(min_row=2, values_only=True):
            student, hw, code_dir, snapshot_file, file_size, file_mtime = row
            
            # 학생이 해당 행에 해당하는지 확인
            if student != student_dir:
                continue
            
            # 파일명에서 타임스탬프만 추출 (파일명이 타임스탬프 형식으로 되어 있다고 가정)
            timestamp_str = ''.join(filter(str.isdigit, snapshot_file))  # 숫자만 추출
            if timestamp_str:  # 타임스탬프가 존재하면
                time = datetime.fromtimestamp(int(timestamp_str))
            else:
                continue  # 타임스탬프가 없으면 건너뜀
            
            # 데이터 추가
            times.append(time)
            file_sizes.append(file_size)
        
        # 시간과 파일 크기를 한 쌍으로 묶어서 정렬
        sorted_times_file_sizes = sorted(zip(times, file_sizes), key=lambda x: x[0])
        sorted_times, sorted_file_sizes = zip(*sorted_times_file_sizes)  # 다시 분리
        
        # 그래프 그리기
        plt.figure(figsize=(10, 6))
        plt.plot(sorted_times, sorted_file_sizes, marker='o', label=f"{student_dir}")
        plt.xlabel('Time')
        plt.ylabel('File Size (bytes)')
        plt.title(f"{hw_dir} - {student_dir}")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.legend()
        
        # 학생별 그래프 저장
        output_filename = os.path.join(output_folder, f"{student_dir}_{hw_dir}")
        plt.savefig(output_filename)  # 그래프를 PNG 파일로 저장
        plt.close()  # 그래프를 닫아 메모리 절약
        print(f"{student_dir}_{hw_dir}")
