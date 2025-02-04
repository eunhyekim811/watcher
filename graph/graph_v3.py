import os
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.cm as cm
import numpy as np
from matplotlib.dates import DateFormatter

def extract_timestamp_from_filename(filename):
    # 파일명에서 숫자만 추출
    numbers = re.findall(r'\d+', filename)
    
    # 10자리 숫자(타임스탬프)만 처리
    for num in numbers:
        if len(num) == 10:  # Unix 타임스탬프는 10자리
            try:
                timestamp = int(num)
                return datetime.fromtimestamp(timestamp)
            except ValueError:
                continue
    return None

# 엑셀 파일 로드
wb = load_workbook("/home/ubuntu/watcher/outputs/hw12.xlsx")
output_folder = "./graphs_hw12"

# 그래프 출력 폴더 생성
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 데이터 저장용 딕셔너리 (학생 → 코드 → [(시간, 크기)])
data = {}

# 모든 시트(학생)에서 데이터 읽기
for sheet in wb.sheetnames:
    ws = wb[sheet]  # 현재 시트 (학생)
    student = sheet
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # 헤더 제외
        code_dir, snapshot_file, file_size, file_mtime, timestamp = row
        
        # 스냅샷 파일명에서 시간 추출
        snapshot_time = extract_timestamp_from_filename(snapshot_file)
        if snapshot_time:  # 타임스탬프가 성공적으로 추출된 경우만 처리
            data.setdefault(student, {}).setdefault(code_dir, []).append((snapshot_time, file_size))

# 그래프 그리기
for student, code_dict in data.items():
    plt.figure(figsize=(50, 35))
    
    cmap = plt.colormaps["tab20"]
    colors = cmap(np.arange(len(code_dict)))
    
    # 모든 시간을 하나의 리스트로 모으기
    all_times = []
    for code, snapshots in code_dict.items():
        snapshots.sort()
        times, _ = zip(*snapshots)
        all_times.extend(times)
    
    # 중복 제거하고 정렬
    all_times = sorted(set(all_times))
    formatted_times = [t.strftime("%m/%d %H:%M:%S") for t in all_times]
    
    # 시간을 인덱스로 매핑하는 딕셔너리 생성
    time_to_idx = {t: i for i, t in enumerate(all_times)}
    
    for idx, (code, snapshots) in enumerate(code_dict.items()):
        snapshots.sort()
        times, file_sizes = zip(*snapshots)
        
        # 각 시간을 인덱스로 변환
        time_indices = [time_to_idx[t] for t in times]
        
        plt.plot(time_indices, file_sizes, marker='o', markersize=9, 
                label=os.path.basename(code), color=colors[idx])

    # x축 설정
    plt.xticks(range(len(formatted_times)), 
               formatted_times,
               rotation=90,
               ha='center',
               va='top',
               fontsize=5)
    
    # 그래프 여백 조정
    plt.subplots_adjust(bottom=0.3)
    
    plt.yticks(fontsize=16)
    plt.xlabel("Time", fontsize=5)
    plt.ylabel("File Size (bytes)", fontsize=20)
    plt.title(f"HW12 - {student}", fontsize=24, pad=20)
    
    plt.legend(title="Code Files", loc='upper left', 
              fontsize='30', frameon=True)
    
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.tight_layout()

    # 그래프 저장
    output_filename = os.path.join(output_folder, f"{student}.png")
    plt.savefig(output_filename, dpi=300)
    plt.close()
    print(f"저장된 그래프: {output_filename}")
